#!/usr/bin/env python
# coding: utf-8

import polars as pl
import os
from io import BytesIO
#from unidecode import unidecode

import base64
import datetime
import pytz

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import dash
#from dash import Dash, dcc, html, dash_table, Input, Output, State, callback, _dash_renderer
from dash import dcc
import dash_ag_grid as dag
#import dash_daq as daq
import dash_bootstrap_components as dbc
from dash_extensions.enrich import Input, Output, State, DashProxy, MultiplexerTransform, html
from dash.exceptions import PreventUpdate

from collections import namedtuple
import pickle

from sqlalchemy import create_engine, URL, text, insert
from sqlalchemy.orm import Session
from sqlalchemy.ext.automap import automap_base

### Conexión a la base de datos
# parámetros de conexión

#objeto_url = URL.create(
#    'postgresql+psycopg2',
#    username = os.environ['PGUSER'],
#    password = os.environ['PGPASSWORD'],
#    host = os.environ['PGHOST'],
#    port = os.environ['PGPORT'],
#    database = os.environ['PGDATABASE'],
#)

#engine = create_engine(objeto_url)
engine = create_engine(os.environ['DATABASE_URL'])

### Fuentes externas
Conversion = namedtuple('Conversion', ['app', 'ex', 'mig', 'lec', 'mod'])

with open('./data/conversion.pkl', 'rb') as f:
    lista_conv = pickle.load(f)

dic_excel = {it.app: it.ex for it in lista_conv if it.ex != None}
dic_migra = {it.app: it.mig for it in lista_conv if it.mig != None}

lista_total = [it.app for it in lista_conv]
lista_app = [it.app for it in lista_conv if it.lec == 1]
lista_mod = [it.app for it in lista_conv if it.mod == 1]
lista_excel = list(dic_excel.keys())
lista_migra = list(dic_migra.keys())

columnas = list(dic_excel.values())

pais = dict(
    pl.read_excel(
        './data/catalogo_pais.xlsx',
        sheet_name='Catálogo_país',
        columns='A:D',
    )
    .select(['codigo', 'pais_es'])
    .cast({'codigo': pl.Int16})
    .rows()
)

comunas = (
    pl.read_excel('./data/cod_regiones.xlsx')
    .get_column('Nombre_Comuna')
    .to_list()
)

cambio_nombre = {x.ex: x.app for x in lista_conv if x.ex != None}


### Diccionarios
documento = {'P': 'Pasaporte', 'R': 'RUN'}
digito = [str(x) for x in range(10)] + ['K']
sexo = {'M': 'Mujer', 'H': 'Hombre', 'NB': 'No Binario'}
tipo_residencia = {1: 'Extranjero CON residencia previa en el país', 2: 'Extranjero SIN residencia previa en el país'}
tipo_programa = {1: 'Programas de intercambio', 2: 'Pasantías médicas', 3: 'Cursos especiales', 4: 'Otros'}
jornada = {1: 'Diurno', 2: 'Vespertino', 3: 'Otro'}
convenio = {1: 'Sí', 2: 'No'}
informacion = {1: 'Definitiva', 2: 'Preliminar'}
estatus = {1: 'Vigente', 2: 'Cancelada', 3: 'Desistida'}

dias = [x for x in range(1, 32)]
meses = dict(zip([x for x in range(1, 13)], ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']))
anos = [x for x in range(1980, 2011)]

universidades = {
    1: 'Universidad Gabriela Mistral',
    2: 'Universidad Finis Terrae',
    4: 'Universidad Central de Chile',
    9: 'Universidad del Alba',
    11: 'Universidad Academia de Humanismo Cristiano',
    13: 'Universidad Santo Tomás',
    17: 'Universidad SEK',
    19: 'Universidad de Las Américas',
    22: 'Universidad de Viña del Mar',
    26: 'Universidad de Artes, Ciencias y Comunicación UNIACC',
    31: 'Universidad Autónoma de Chile',
    38: 'Universidad Adventista de Chile',
    39: 'Universidad San Sebastián',
    42: 'Universidad Católica Cardenal Raúl Silva Henríquez',
    50: "Universidad Bernardo O'Higgins",
    68: 'Universidad Miguel de Cervantes'
}


### Esquemas
# esquema de lectura de datos desde la base de datos (49 variables)
sch_base = {
    'id': pl.Int16,
    'documento': pl.Utf8,
    'numero': pl.Int64,
    'dv': pl.Utf8,
    'apellido_1': pl.Utf8,
    'apellido_1_u': pl.Utf8,
    'apellido_2': pl.Utf8,
    'apellido_2_u': pl.Utf8,
    'nombres': pl.Utf8,
    'nombres_u': pl.Utf8,
    'sexo': pl.Utf8,
    'fecha_nac': pl.Date,
    'dia_nac': pl.Int8,
    'mes_nac': pl.Int8,
    'ano_nac': pl.Int16,
    'nacional': pl.Int16,
    'pais': pl.Int16,
    'tipo_residencia': pl.Int8,
    'programa': pl.Utf8,
    'programa_u': pl.Utf8,
    'cod_programa': pl.Utf8,
    'tipo_programa': pl.Int8,
    'esp_programa': pl.Utf8,
    'esp_programa_u': pl.Utf8,
    'duracion': pl.Int8,
    'inicio_programa': pl.Date,
    'dia_ini': pl.Int8,
    'mes_ini': pl.Int8,
    'ano_ini': pl.Int16,
    'termino_programa': pl.Date,
    'dia_ter': pl.Int8,
    'mes_ter': pl.Int8,
    'ano_ter': pl.Int16,
    'jornada': pl.Int8,
    'comuna': pl.Utf8,
    'comuna_u': pl.Utf8,
    'universidad_origen': pl.Utf8,
    'universidad_origen_u': pl.Utf8,
    'pais_universidad': pl.Int16,
    'convenio': pl.Int8,
    'fecha_postulacion': pl.Date,
    'dia_pos': pl.Int8,
    'mes_pos': pl.Int8,
    'ano_pos': pl.Int16,
    'condicion_inf': pl.Int8,
    'estatus': pl.Int8,
    'creada': pl.Datetime('us', 'America/Santiago'),
    'universidad': pl.Int8,
    'validada': pl.Int8,
}

sch_base_48 = sch_base.copy()
sch_base_48.pop('id')

# esquema de lectura de datos almacenados localmente
sch_base_loc = sch_base.copy()
sch_base_loc.update(
    {
        'fecha_nac': pl.Utf8,
        'inicio_programa': pl.Utf8,
        'termino_programa': pl.Utf8,
        'fecha_postulacion': pl.Utf8,
        'creada': pl.Utf8,
    }
)

# esquema de lectura de datos desde la aplicación (34 variables)
dic_sch_app = {
    'documento': pl.Utf8,
    'numero': pl.Int64,
    'dv': pl.Utf8,
    'apellido_1': pl.Utf8,
    'apellido_2': pl.Utf8,
    'nombres': pl.Utf8,
    'sexo': pl.Utf8,
    'dia_nac': pl.Int8,
    'mes_nac': pl.Int8,
    'ano_nac': pl.Int16,
    'nacional': pl.Int16,
    'pais': pl.Int16,
    'tipo_residencia': pl.Int8,
    'programa': pl.Utf8,
    'cod_programa': pl.Utf8,
    'tipo_programa': pl.Int8,
    'esp_programa': pl.Utf8,
    'duracion': pl.Int8,
    'dia_ini': pl.Int8,
    'mes_ini': pl.Int8,
    'ano_ini': pl.Int16,
    'dia_ter': pl.Int8,
    'mes_ter': pl.Int8,
    'ano_ter': pl.Int16,
    'jornada': pl.Int8,
    'comuna': pl.Utf8,
    'universidad_origen': pl.Utf8,
    'pais_universidad': pl.Int16,
    'convenio': pl.Int8,
    'dia_pos': pl.Int8,
    'mes_pos': pl.Int8,
    'ano_pos': pl.Int16,
    'condicion_inf': pl.Int8,
    'estatus': pl.Int8,
}

sch_app = pl.Schema(dic_sch_app)

# esquema lectura de datos desde Excel (26 variables)
dic_sch_excel = {
    'documento': pl.Utf8,
    'numero': pl.Int64,
    'dv': pl.Utf8,
    'apellido_1': pl.Utf8,
    'apellido_2': pl.Utf8,
    'nombres': pl.Utf8,
    'sexo': pl.Utf8,
    'fecha_nac': pl.Date,
    'nacional': pl.Utf8,
    'pais': pl.Utf8,
    'tipo_residencia': pl.Utf8,
    'programa': pl.Utf8,
    'cod_programa': pl.Utf8,
    'tipo_programa': pl.Utf8,
    'esp_programa': pl.Utf8,
    'duracion': pl.Int8,
    'inicio_programa': pl.Date,
    'termino_programa': pl.Date,
    'jornada': pl.Utf8,
    'comuna': pl.Utf8,
    'universidad_origen': pl.Utf8,
    'pais_universidad': pl.Utf8,
    'convenio': pl.Utf8,
    'fecha_postulacion': pl.Date,
    'condicion_inf': pl.Utf8,
    'estatus': pl.Utf8,
}

sch_excel = pl.Schema(dic_sch_excel)

# cración de namedtuples
Registro = namedtuple('Registro', lista_total)
Registro_48 = namedtuple('Registro_48', lista_total[1:])  # 48 variables
Registro_excel = namedtuple('Registro_excel', lista_excel)
Registro_app = namedtuple('Registro_app', lista_app)
Registro_mod = namedtuple('Registro_mod', lista_mod)


### Funciones

inv = lambda x: {v: k for k, v in x.items()}

tz = pytz.timezone('America/Santiago')

#### Funciones de transformación de datos
# aplica a la lectura de datos desde excel y la aplicación

def elimina_espacios(df):
    return (
        df
        .with_columns(
            pl.col('apellido_1').str.strip_chars(),
            pl.col('apellido_2').str.strip_chars(),
            pl.col('nombres').str.strip_chars(),
            pl.col('programa').str.strip_chars(),
            pl.col('esp_programa').str.strip_chars(),
            pl.col('universidad_origen').str.strip_chars(),
        )
    )


def elimina_duplicados(df, variables):
    return df.unique(subset=variables, keep='first', maintain_order=True)


def verifica_duplicados(df):
    return (
        df
        .select(lista_app)
        .is_duplicated()
        .any()
    )


def elimina_no_definidos(df):
    return df.filter(pl.col('numero').is_not_null() & pl.col('programa').is_not_null())


def selec_ultimo(df):
    return (
        df
        .group_by(['numero', 'programa'], maintain_order=True).last()
        .select(lista_excel)
    )


mapa_acentos = {
    'á': 'a',
    'é': 'e',
    'í': 'i',
    'ó': 'o',
    'ú': 'u',
    'Á': 'A',
    'É': 'E',
    'Í': 'I',
    'Ó': 'O',
    'Ú': 'U',
}


def mayusc(item):
    return (''.join([mapa_acentos.get(i, i) for i in list(item)])).upper()


# convierte excel a base de trabajo
def convierte_excel(df, usuario):
    return (
        df
        .with_columns(
            pl.col('documento').replace(inv(documento)),
            pl.col('apellido_1').map_elements(lambda x: mayusc(x), return_dtype=pl.Utf8).alias('apellido_1_u'),
            pl.col('apellido_2').map_elements(lambda x: mayusc(x), return_dtype=pl.Utf8).alias('apellido_2_u'),
            pl.col('nombres').map_elements(lambda x: mayusc(x), return_dtype=pl.Utf8).alias('nombres_u'),
            pl.col('sexo').replace(inv(sexo)),
            pl.col('fecha_nac').dt.day().alias('dia_nac').cast(pl.Int8),
            pl.col('fecha_nac').dt.month().alias('mes_nac').cast(pl.Int8),
            pl.col('fecha_nac').dt.year().alias('ano_nac').cast(pl.Int16),
            pl.col('nacional').replace_strict(inv(pais), return_dtype=pl.Int16),
            pl.col('pais').replace_strict(inv(pais), return_dtype=pl.Int16),
            pl.col('tipo_residencia').replace_strict(inv(tipo_residencia), return_dtype=pl.Int8),
            pl.col('programa').map_elements(lambda x: mayusc(x), return_dtype=pl.Utf8).alias('programa_u'),
            pl.col('tipo_programa').replace_strict(inv(tipo_programa), return_dtype=pl.Int8),
            pl.col('esp_programa').map_elements(lambda x: mayusc(x), return_dtype=pl.Utf8).alias('esp_programa_u'),
            pl.col('inicio_programa').dt.day().alias('dia_ini').cast(pl.Int8),
            pl.col('inicio_programa').dt.month().alias('mes_ini').cast(pl.Int8),
            pl.col('inicio_programa').dt.year().alias('ano_ini').cast(pl.Int16),
            pl.col('termino_programa').dt.day().alias('dia_ter').cast(pl.Int8),
            pl.col('termino_programa').dt.month().alias('mes_ter').cast(pl.Int8),
            pl.col('termino_programa').dt.year().alias('ano_ter').cast(pl.Int16),
            pl.col('jornada').replace_strict(inv(jornada), return_dtype=pl.Int8),
            pl.col('comuna').map_elements(lambda x: mayusc(x), return_dtype=pl.Utf8).alias('comuna_u'),
            pl.col('universidad_origen').map_elements(lambda x: mayusc(x), return_dtype=pl.Utf8).alias('universidad_origen_u'),
            pl.col('pais_universidad').replace_strict(inv(pais), return_dtype=pl.Int16),
            pl.col('convenio').replace_strict(inv(convenio), return_dtype=pl.Int8),
            pl.col('fecha_postulacion').dt.day().alias('dia_pos').cast(pl.Int8),
            pl.col('fecha_postulacion').dt.month().alias('mes_pos').cast(pl.Int8),
            pl.col('fecha_postulacion').dt.year().alias('ano_pos').cast(pl.Int16),
            pl.col('condicion_inf').replace_strict(inv(informacion), return_dtype=pl.Int8),
            pl.col('estatus').replace_strict(inv(estatus), return_dtype=pl.Int8),
            pl.lit(usuario).alias('universidad').cast(pl.Int8),
            pl.lit(datetime.datetime.now(tz)).dt.round('1s').alias('creada').cast(pl.Datetime('us', 'America/Santiago')),
            pl.lit(0).alias('validada').cast(pl.Int8),
        )
    )


def back_excel(df):
    lista = lista_excel.copy()
    if 'sep' in df.columns:
        lista += ['sep']
    if 'id' in df.columns:
        lista += ['id', 'universidad']
    return (
        df
        .select(lista)
        .with_columns(
            pl.col('documento').replace(documento),
            pl.col('sexo').replace(sexo),
            pl.col('nacional').replace_strict(pais, return_dtype=pl.Utf8),
            pl.col('pais').replace_strict(pais, return_dtype=pl.Utf8),
            pl.col('tipo_residencia').replace_strict(tipo_residencia, return_dtype=pl.Utf8),
            pl.col('tipo_programa').replace_strict(tipo_programa, return_dtype=pl.Utf8),
            pl.col('jornada').replace_strict(jornada, return_dtype=pl.Utf8),
            pl.col('pais_universidad').replace_strict(pais, return_dtype=pl.Utf8),
            pl.col('convenio').replace_strict(convenio, return_dtype=pl.Utf8),
            pl.col('condicion_inf').replace_strict(informacion, return_dtype=pl.Utf8),
            pl.col('estatus').replace_strict(estatus, return_dtype=pl.Utf8),
        )
    )


# función que convierte ingreso via app en df y tupla
def fn_fecha(dia, mes, ano):
    if (dia != None) & (mes != None) & (ano != None):
        return datetime.date(ano, mes, dia)
    else:
        return None
    

round_time = lambda dt: (dt + datetime.timedelta(milliseconds=500)).replace(microsecond=0)


def convierte_app(ntp, usuario):
    dic = ntp._asdict()

    dic['fecha_nac'] = fn_fecha(ntp.dia_nac, ntp.mes_nac, ntp.ano_nac)
    dic['inicio_programa'] = fn_fecha(ntp.dia_ini, ntp.mes_ini, ntp.ano_ini)
    dic['termino_programa'] = fn_fecha(ntp.dia_ter, ntp.mes_ter, ntp.ano_ter)
    dic['fecha_postulacion'] = fn_fecha(ntp.dia_pos, ntp.mes_pos, ntp.ano_pos)

    dic['apellido_1_u'] = mayusc(ntp.apellido_1) if ntp.apellido_1 != None else None
    dic['apellido_2_u'] = mayusc(ntp.apellido_2) if ntp.apellido_2 != None else None
    dic['nombres_u'] = mayusc(ntp.nombres) if ntp.nombres != None else None
    dic['programa_u'] = mayusc(ntp.programa) if ntp.programa != None else None
    dic['esp_programa_u'] = mayusc(ntp.esp_programa) if ntp.esp_programa != None else None
    dic['comuna_u'] = mayusc(ntp.comuna) if ntp.comuna != None else None
    dic['universidad_origen_u'] = mayusc(ntp.universidad_origen) if ntp.universidad_origen != None else None

    dic['universidad'] = usuario
    dic['creada'] = round_time(datetime.datetime.now(tz))
    dic['validada'] = 1  # FUNCIÓN QUE DETERMINA LA CONDICIÓN DE VALIDADA

    df = pl.DataFrame(dic, schema_overrides=sch_base_48)
    tupla = Registro_48(**dic)  # 48 variables
    
    return df, tupla, ntp.numero


def convierte_mod(ntp, resto):
    dic = ntp._asdict() | resto

    dic['fecha_nac'] = fn_fecha(ntp.dia_nac, ntp.mes_nac, ntp.ano_nac)
    dic['inicio_programa'] = fn_fecha(ntp.dia_ini, ntp.mes_ini, ntp.ano_ini)
    dic['termino_programa'] = fn_fecha(ntp.dia_ter, ntp.mes_ter, ntp.ano_ter)
    dic['fecha_postulacion'] = fn_fecha(ntp.dia_pos, ntp.mes_pos, ntp.ano_pos)

    dic['apellido_1_u'] = mayusc(ntp.apellido_1) if ntp.apellido_1 != None else None
    dic['apellido_2_u'] = mayusc(ntp.apellido_2) if ntp.apellido_2 != None else None
    dic['nombres_u'] = mayusc(ntp.nombres) if ntp.nombres != None else None
    dic['programa_u'] = mayusc(dic['programa']) if dic['programa'] != None else None
    dic['esp_programa_u'] = mayusc(ntp.esp_programa) if ntp.esp_programa != None else None
    dic['comuna_u'] = mayusc(ntp.comuna) if ntp.comuna != None else None
    dic['universidad_origen_u'] = mayusc(ntp.universidad_origen) if ntp.universidad_origen != None else None

    dic['creada'] = round_time(datetime.datetime.now(tz))
    dic['validada'] = 1

    tupla = Registro(**dic)
    
    return tupla


#### Funciones de lectura
# función que convierte columnas datetime a str y vice versa

def convierte_fecha_str(df):
    return (df
        .with_columns([
            pl.col('fecha_nac').dt.strftime('%Y-%m-%d').cast(pl.Utf8),
            pl.col('inicio_programa').dt.strftime('%Y-%m-%d').cast(pl.Utf8),
            pl.col('termino_programa').dt.strftime('%Y-%m-%d').cast(pl.Utf8),
            pl.col('fecha_postulacion').dt.strftime('%Y-%m-%d').cast(pl.Utf8),
            pl.col('creada').dt.strftime('%Y-%m-%d %H:%M:%S').cast(pl.Utf8),
        ])
    )


def convierte_str_fecha(df):
    return (df
        .with_columns([
            pl.col('fecha_nac').str.to_datetime(format='%Y-%m-%d').dt.date().cast(pl.Date),
            pl.col('inicio_programa').str.to_datetime(format='%Y-%m-%d').dt.date().cast(pl.Date),
            pl.col('termino_programa').str.to_datetime(format='%Y-%m-%d').dt.date().cast(pl.Date),
            pl.col('fecha_postulacion').str.to_datetime(format='%Y-%m-%d').dt.date().cast(pl.Date),
            pl.col('creada').str.to_datetime(format='%Y-%m-%d %H:%M:%S').cast(pl.Datetime()), #('us', 'America/Santiago')),
        ])
    )


# datos de solicitudes de visa: selección por usuario
def lectura(usuario):
    df = (
        (pl.read_database(
            query = f"SELECT * FROM visas WHERE universidad = {usuario} ORDER BY id", 
            connection = engine,
        )
        .cast(sch_base))
    )

    return df, df.to_dicts()
    

def lectura_conv(usuario):
    df = convierte_fecha_str(
        (pl.read_database(
            query = f"SELECT * FROM visas WHERE universidad = {usuario} ORDER BY id",
            connection = engine,
        )
        .cast(sch_base))
    )

    return df, df.to_dicts()


### Constructor
# objeto SqlAlchemy

Base = automap_base()
Base.prepare(autoload_with=engine)

Visa = Base.classes.visas

# para ingreso unitario de información
def obj_visas(tup):
    return Visa(
        documento = tup.documento,
        numero = tup.numero,
        dv = tup.dv,
        apellido_1 = tup.apellido_1,
        apellido_1_u = tup.apellido_1_u,
        apellido_2 = tup.apellido_2,
        apellido_2_u = tup.apellido_2_u,
        nombres = tup.nombres,
        nombres_u = tup.nombres_u,
        sexo = tup.sexo,
        fecha_nac = tup.fecha_nac,
        dia_nac = tup.dia_nac,
        mes_nac = tup.mes_nac,
        ano_nac = tup.ano_nac,
        nacional = tup.nacional,
        pais = tup.pais,
        tipo_residencia = tup.tipo_residencia,
        programa = tup.programa,
        programa_u = tup.programa_u,
        cod_programa = tup.cod_programa,
        tipo_programa = tup.tipo_programa,
        esp_programa = tup.esp_programa,
        esp_programa_u = tup.esp_programa_u,
        duracion = tup.duracion,
        inicio_programa = tup.inicio_programa,
        dia_ini = tup.dia_ini,
        mes_ini = tup.mes_ini,
        ano_ini = tup.ano_ini,
        termino_programa = tup.termino_programa,
        dia_ter = tup.dia_ter,
        mes_ter = tup.mes_ter,
        ano_ter = tup.ano_ter,
        jornada = tup.jornada,
        comuna = tup.comuna,
        comuna_u = tup.comuna_u,
        universidad_origen = tup.universidad_origen,
        universidad_origen_u = tup.universidad_origen_u,
        pais_universidad = tup.pais_universidad,
        convenio = tup.convenio,
        fecha_postulacion = tup.fecha_postulacion,
        dia_pos = tup.dia_pos,
        mes_pos = tup.mes_pos,
        ano_pos = tup.ano_pos,
        condicion_inf = tup.condicion_inf,
        estatus = tup.estatus,
        creada = tup.creada,
        universidad = tup.universidad,
        validada = tup.validada,
    )


### Funciones
# ingreso en bloque de información

def agrega_excel(df): #, usuario):
    with Session(engine) as session:
        session.execute(
            insert(Visa),
            df.to_dicts(),
        )
        session.commit()

#    return lectura_conv(usuario)[1]  # para almacenamiento local


# In[51]:


# ingreso individual

def nueva_visas(tup): #, usuario):
    visa = obj_visas(tup)

    with Session(engine) as session:
        session.add(visa)
        session.commit()

#    return lectura_conv(usuario)[1]  # para almacenamiento local


def elimina_visas(id, usuario):

    with Session(engine) as session:
        elimina = session.query(Visa).filter(Visa.id == id).first()
        session.delete(elimina)
        session.commit()

    return lectura_conv(usuario)[1]  # se mantiene hasta evaluar


def modifica_visas(tup): #, usuario):
    id_loc = tup.id

    with Session(engine) as session:
        modifica = session.query(Visa).filter(Visa.id == id_loc).first()
    
        modifica.documento = tup.documento,
        modifica.dv = tup.dv,
        modifica.apellido_1 = tup.apellido_1,
        modifica.apellido_1_u = tup.apellido_1_u,
        modifica.apellido_2 = tup.apellido_2,
        modifica.apellido_2_u = tup.apellido_2_u,
        modifica.nombres = tup.nombres,
        modifica.nombres_u = tup.nombres_u,
        modifica.sexo = tup.sexo,
        modifica.fecha_nac = tup.fecha_nac,
        modifica.dia_nac = tup.dia_nac,
        modifica.mes_nac = tup.mes_nac,
        modifica.ano_nac = tup.ano_nac,
        modifica.nacional = tup.nacional,
        modifica.pais = tup.pais,
        modifica.tipo_residencia = tup.tipo_residencia,
        modifica.cod_programa = tup.cod_programa,
        modifica.tipo_programa = tup.tipo_programa,
        modifica.esp_programa = tup.esp_programa,
        modifica.esp_programa_u = tup.esp_programa_u,
        modifica.duracion = tup.duracion,
        modifica.inicio_programa = tup.inicio_programa,
        modifica.dia_ini = tup.dia_ini,
        modifica.mes_ini = tup.mes_ini,
        modifica.ano_ini = tup.ano_ini,
        modifica.termino_programa = tup.termino_programa,
        modifica.dia_ter = tup.dia_ter,
        modifica.mes_ter = tup.mes_ter,
        modifica.ano_ter = tup.ano_ter,
        modifica.jornada = tup.jornada,
        modifica.comuna = tup.comuna,
        modifica.comuna_u = tup.comuna_u,
        modifica.universidad_origen = tup.universidad_origen,
        modifica.universidad_origen_u = tup.universidad_origen_u,
        modifica.pais_universidad = tup.pais_universidad,
        modifica.convenio = tup.convenio,
        modifica.fecha_postulacion = tup.fecha_postulacion,
        modifica.dia_pos = tup.dia_pos,
        modifica.mes_pos = tup.mes_pos,
        modifica.ano_pos = tup.ano_pos,
        modifica.condicion_inf = tup.condicion_inf,
        modifica.estatus = tup.estatus,
        modifica.creada = tup.creada,
        modifica.universidad = tup.universidad,
        modifica.validada = tup.validada,
        session.commit()

#    return lectura_conv(usuario)[1]  # para almacenamiento local


# mensaje
dic_causas = {
    1: 'El archivo ingresado no tiene la extensión de un archivo Excel.',
    2: 'El archivo ingresado no es reconocido como Excel.',
    3: 'El archivo ingresado no tiene una Hoja denominada "Postulante".',
    4: 'El archivo ingresado no tiene las columnas esperadas.',
    5: 'El archivo ingresado está vacío.',
    6: 'No existen nuevos registros o modificaciones a registros que añadir.',
}

def mensaje(causa):
    return html.Div([
        dbc.Row(
            dbc.Col([
                dbc.Alert(dic_causas[causa], color='primary', style={'textAlign': 'center', 'marginTop': '25px'}),
            ],
            width={"size": 8, "offset": 2},
            ),
        ),
        dbc.Row(
            dbc.Button('Continuar', id='btn-restituye', n_clicks=0, outline=True, color='primary', style={'width': '12%'}),
            justify='center'
        ),
    ])


agrega_df = lambda df, i: df.with_columns(pl.lit(i).alias('df').cast(pl.Int8))


def clasifica(cantidad, dup):
    if dup == 0:
        if cantidad == 1:
            return 1  # nuevo
        else:
            return 2  # actualiza
    else:
        return 0 


# función que diferencia duplicados, nuevos y actualizaciones
def separa(nueva, usuario, omite_dup=True):
    criterio = (pl.col('df') == 1)
    if omite_dup:
        criterio &= (pl.col('sep') != 0)
    base = agrega_df(lectura(usuario)[0], 0)
    df_con = pl.concat([base, nueva], how='diagonal')
    df_loc = (
        df_con
        .sort(['numero', 'programa', 'creada'])
        .with_columns(
            id = pl.col('id').forward_fill().backward_fill().over(['numero', 'programa']),
            cuenta = pl.len().over(['numero', 'programa']),
            dup = pl.struct(lista_excel).is_duplicated().cast(pl.Int8),
        )
        .with_columns(
            pl.struct(['cuenta', 'dup']).map_elements(lambda x: clasifica(x['cuenta'], x['dup']), return_dtype=pl.Int8).alias('sep')
        )
        .filter(criterio)
        .drop(['df', 'cuenta', 'dup'])
    )

    return df_loc


columns_def = [
    {'field': 'documento', 'headerName': 'Documento', 'width': 120},
    {'field': 'numero', 'headerName': 'Número', 'width': 150},
    {'field': 'dv', 'headerName': 'DV', 'width': 60},
    {'field': 'apellido_1', 'headerName': 'Apellido 1'},
    {'field': 'apellido_2', 'headerName': 'Apellido 2'},
    {'field': 'nombres', 'headerName': 'Nombres', 'width': 60},
    {'field': 'sexo', 'headerName': 'Sexo'},
    {'field': 'fecha_nac', 'headerName': 'Fecha de nacimiento'},
    {'field': 'pais', 'headerName': 'Nacionalidad'},
    {'field': 'nacional', 'headerName': 'País'},
    {'field': 'tipo_residencia', 'headerName': 'Tipo de residencia'},
    {'field': 'programa', 'headerName': 'Nombre programa'},
    {'field': 'cod_programa', 'headerName': 'Código del programa'},
    {'field': 'tipo_programa', 'headerName': 'Tipo programa'},
    {'field': 'esp_programa', 'headerName': 'Especificación programa'},
    {'field': 'duracion', 'headerName': 'Duración'},
    {'field': 'inicio_programa', 'headerName': 'Fecha de inicio'},
    {'field': 'termino_programa', 'headerName': 'Fecha de término'},
    {'field': 'jornada', 'headerName': 'Jornada'},
    {'field': 'comuna', 'headerName': 'Comuna'},
    {'field': 'universidad_origen', 'headerName': 'Universidad de origen'},
    {'field': 'pais_universidad', 'headerName': 'País de la universidad'},
    {'field': 'convenio', 'headerName': 'Existe convenio'},
    {'field': 'fecha_postulacion', 'headerName': 'Fecha postulación'},
    {'field': 'condicion_inf', 'headerName': 'Condición información'},
    {'field': 'estatus', 'headerName': 'Estatus postulación'},
    {'field': 'sep', 'hide': True},
]

getRowStyle = {
    "styleConditions": [
        {
            "condition": "params.data['sep'] == 1",  # nuevo
            "style": {"backgroundColor": "#ecffd8"},
        },
    ],
    "defaultStyle": {"backgroundColor": "#ffffd8"},  # actualización
}

boton_ingresa_excel = dbc.Col([
    dbc.Row([
        html.Button('Cancelar', id='btn-canc-inf-ex', n_clicks=0, className='btn btn-outline-primary', style={'width': '30%'}),
        html.Button('Ingreso de información', id='btn-ing-inf-ex', n_clicks=0, className='btn btn-outline-primary', style={'width': '30%', 'marginLeft': '10px'}),
    ],
        justify='end',
        style={'margin': '10px 0px 0px'}
    ),
    dbc.Row(
        dcc.RadioItems(
            id='sel-registros',
            options=[
               {'label': 'Todos', 'value': 1},
               {'label': 'Solo nuevos', 'value': 0},
            ],
            value = 1,
            inline=True,
            style = {'textAlign': 'start', 'width': '32%', 'display': 'inline-block'},
            labelStyle = {'display': 'inline-block', 'fontSize': '16px', 'fontWeight': 'normal'},
            inputStyle = {'marginRight': '5px', 'marginLeft': '20px'},
        ),
        justify='end',
    )],
width=7
)


def parse_excel(contenido, filename, usuario):
    contenido_type, contenido_string = contenido.split(',')

    decoded = base64.b64decode(contenido_string)

    if 'xls' in filename:
        try:
            wb = load_workbook(filename=BytesIO(decoded))
            if 'Postulante' in wb.sheetnames:
                nueva = pl.read_excel(
                    BytesIO(decoded),
                    sheet_name='Postulante',
                    read_options={'header_row': 1},
                )
                if columnas == nueva.columns:
                    if nueva.is_empty():
                        return (mensaje(5), dash.no_update)
                    else:
                        nueva = nueva.rename(cambio_nombre).cast(sch_excel)
                        nueva = elimina_espacios(nueva)
                        nueva = elimina_duplicados(nueva, lista_excel)
                        nueva = elimina_no_definidos(nueva)
                        nueva = selec_ultimo(nueva)
                        if nueva.is_empty():
                            return (mensaje(6), dash.no_update)
                        else:
                            nueva = convierte_excel(nueva, usuario)
                            nueva = agrega_df(nueva, 1)
                            nueva = separa(nueva, usuario)
                            if nueva.is_empty():
                                return (mensaje(6), dash.no_update)
                else:
                    return (mensaje(4), dash.no_update)
            else:
                return (mensaje(3), dash.no_update)
        except InvalidFileException:
            return (mensaje(2), dash.no_update)
    else:
        return (mensaje(1), dash.no_update)
    
    return (
        html.Div([
            html.H4('Registros nuevos o que modifican registros existentes'),
            html.P(f'Archivo leído: {filename}', style={'fontSize': 14, 'marginLeft': '20px'}),
            dag.AgGrid(
                rowData=back_excel(nueva).to_dicts(),
                columnDefs=columns_def,
                dashGridOptions = {'domLayout': 'autoHeight'},
                getRowStyle=getRowStyle,
                defaultColDef = {'editable': False}
            ),
            dbc.Row([
                dbc.Col(
                    html.Div([
                        html.Img(src='./assets/verde.png', style={'width': '10%', 'height': '80%'}),
                        html.P('Registros nuevos', style={'size': 16, 'margin': '0px 10px', 'display': 'inline-block'}),
                        html.Img(src='./assets/amarillo.png', style={'width': '10%', 'height': '80%'}),
                        html.P('Modifican registros', style={'size': 16, 'margin': '0px 10px', 'display': 'inline-block'}),
                    ], style={'marginTop': '10px', 'display': 'inline-block'}),
                    width=5
                ),
                boton_ingresa_excel,
            ],
            justify='between',
            )
        ]),
        convierte_fecha_str(nueva).to_dicts()
    )


def agrega_inf_excel(lista, todos):
    df_loc = convierte_str_fecha(pl.DataFrame(lista, schema_overrides=(sch_base_loc | {'sep': pl.Int8})))
    dic_df = df_loc.partition_by('sep', include_key=False, as_dict=True)
    nuevos = dic_df.get((1,), 0)
    actual = dic_df.get((2,), 0)
    if not isinstance(nuevos, int):
        agrega_excel(nuevos.drop('id'))
    if todos and not isinstance(actual, int):
        for dic in actual.to_dicts():
            modifica_visas(Registro(**dic))


def parse_app(tup, usuario):
    nuevo, tup_loc, num = convierte_app(tup, usuario)
    nuevo = agrega_df(nuevo, 1)
    nuevo = separa(nuevo, usuario, omite_dup=False)

    dic = nuevo.to_dicts()[0]
    tipo = dic.pop('sep')

    if tipo == 0:  # duplicado
        return dash.no_update, dash.no_update, True, dash.no_update, dash.no_update
    elif tipo == 1:  # nuevo
        dic.pop('id')
        nueva_visas(Registro_48(**dic))
        return lectura_conv(usuario)[1], True, dash.no_update, dash.no_update, dash.no_update
    elif tipo == 2:  # actualización
        provisional = convierte_fecha_str(nuevo.drop('sep')).to_dicts()
        return dash.no_update, dash.no_update, dash.no_update, True, provisional


### Construcción de la aplicación
azul = '#2FA4E7'
linea = html.Hr(
    id='linea',
    hidden='hidden',
    style={'marginTop': 5, 'marginBottom': 5, 'border': None, 'borderWidth': '1px', 'width': '100%', 'color': azul, 'opacity': 1},
)
espacio = html.Br()


# función que convierte diccionario en formato de opciones para dropdown
def opciones(dic):
    return [{'label': v, 'value': k} for k, v in dic.items()]


#### Acceso
acceso = html.Div(
    dbc.Container([
        dbc.Row(
            dbc.Col(
                html.Img(src='./assets/cup-logo-2.svg', style={'width': '100%', 'height': '100%'}),
                width={"size": 2, "offset": 5},
                style={'marginBottom': 5}
            ),
        ),
        dbc.Row(
            dbc.Col(
                html.H2(['Gestión de Visas de Estudiantes Extranjeros'], style={'textAlign': 'center', 'color': '#7F7F7F', 'margin': 0}),
            )
        ),
        dbc.Row(
            html.Div([
                dbc.Row([
                    dbc.Col(html.P(html.Span('Usuario:'), className='pAcceso'), width=2),
                    dbc.Col(
                        dcc.Dropdown(opciones(universidades), placeholder='Seleccione su universidad', id='sel-u', style={'width': '100%', 'textAling': 'left', 'background': '#f1f1f1', 'border': '0px'}),
                        align='center',
                        width=10,
                    ),
                ]),
                dbc.Row([
                    dbc.Col(html.P(html.Span('Password:'), className='pAcceso'), width=2),                
                    dbc.Col(dbc.Input(placeholder="Ingrese su password", id='inp-pw', type="password", debounce=True, style={'background': '#f1f1f1', 'border': '0px', 'borderRadius': '0px'}), 
                        align='center',
                        width=6,
                    ),
                ])
            ], className='boxAcceso',
            ), justify='center',
        ),
        dbc.Row(
            html.Button('Ingresar', id='btn-ingresar', n_clicks=0, className='btn btn-outline-primary', style={'width': '12%'}),
            justify='center',
        )
    ],
    style={'marginTop': '20%', 'marginBottom': 5}
    ), id='acceso',
)


#### Encabezado
encabezado = html.Div(
    dbc.Row([
        dbc.Col(
            html.Img(src='./assets/cup-logo-2.svg', style={'width': '100%', 'height': '100%'}),
            width=2,
        ),
        dbc.Col([
            html.H2(['Gestión de Visas de Estudiantes Extranjeros'], style={'textAlign': 'center', 'margin': 0}),
            html.H4(['Período 2025'], style={'textAlign': 'center', 'margin': 0}),
        ], width=8),
    ], align='center'),
    id='encabezado',
    hidden='hidden',
    style={'marginTop': 15, 'marginBottom': 5}
)


#### Usuario
def muestra_usuario():
    return html.Div(
        dbc.Row(
            dbc.Col([
                html.P(html.Span(id='sp-usuario', className='divBorder'), className='pUsuario', style={'textAlign': 'center'}),    
            ])
        ),
        id='muestra-usuario',
        hidden='hidden',
    )


#### Menú
dd_menu = html.Div(
    dbc.Row([
        dbc.Col(
            dbc.DropdownMenu(
                label="Menú",
                size="lg",
                children=[
                    dbc.DropdownMenuItem("Resumen", id='tab-resumen', active=True, n_clicks=0),
                    dbc.DropdownMenuItem("Ingreso de información", id='tab-ingresa_1', n_clicks=0),
                    dbc.DropdownMenuItem("Información desde archivo Excel", id='tab-ingresa_2', n_clicks=0),
                    dbc.DropdownMenuItem("Modificación/eliminación de información", id='tab-modifica', n_clicks=0),
                ]
            ),
            width=1,
        ),
        dbc.Col(html.H3(': Resumen', id='ubica', style={'marginLeft': '30px', 'marginTop': '0.2rem'}), align='center', width=10)
    ]),
    id='menu',
    hidden='hidden',
)


#### Resumen
# función que retorna el 'modo-seleccionado' (contenedor) en ventana Resumen

excluye = ['condicion_inf', 'estatus']
dic_migra_corto = {k: v for k, v in dic_migra.items() if k not in excluye}
lista_migra_corta = [x for x in lista_migra if x not in excluye]

dic_modo = {
    1: [dic_excel, lista_excel],
    0: [dic_migra_corto, lista_migra_corta],
}


def prepara_datos(datos, modo):
    df = (
        convierte_str_fecha(pl.DataFrame(datos).cast(sch_base_loc))
        .cast(sch_base)
        .select(dic_modo[modo][1])
    )
    if modo == 1:
        df = back_excel(df)
    df = df.rename(dic_modo[modo][0])
    return df


def modo_resumen(datos, modo):
    df = (
        pl.DataFrame(datos)
        .cast(sch_base_loc)
        .select(dic_modo[modo][1])
    )
    if modo == 1:
        df = back_excel(df)
    return html.Div([
        dag.AgGrid(
            rowData=df.to_dicts(),
            columnDefs=[{'field': i, 'headerName': dic_modo[modo][0][i]} for i in df.columns],
            dashGridOptions = {'domLayout': 'autoHeight'},
            defaultColDef = {'editable': False}
        ),
    ], id='modo-seleccionado')
        

def vista_resumen():
    return html.Div([
        html.H4('Solicitudes a tramitar'),
        dbc.Row([
            html.P('Forma de visualización:', style={'width': '16%', 'fontSize': '17px', 'margin': '0px 0px 0px 20px', 'display': 'inline-block'}),
            dcc.RadioItems(
                id='selector-modo',
                options=[
                   {'label': 'Normal', 'value': 1},
                   {'label': 'Codificada', 'value': 0},
                ],
                value = 1,
                style = {'textAlign': 'start', 'width': '20%', 'display': 'inline-block'},
                labelStyle = {'display': 'inline-block', 'fontSize': '16px', 'fontWeight': 'normal'},
                inputStyle = {'marginRight': '5px', 'marginLeft': '20px'},
            ),
        ],
        align='center',
        style={'marginBottom': '10px'}
        ),
#        html.Div(id='modo-seleccionado'),  # contenedor
    ])


boton_descarga = dbc.Row([
    html.Button('Descargar archivo de trabajo', id='btn-descarga-excel', n_clicks=0, className='btn btn-outline-primary', style={'width': '25%'}),
    dcc.Download(id='descarga-excel'),
    html.Button('Exportar a Excel', id='btn-exporta-excel', n_clicks=0, className='btn btn-outline-primary', style={'width': '17%', 'marginLeft': '10px'}),
    dcc.Download(id='exporta-excel'),
],
    justify='between',
    style={'margin': '10px 0px 35px'}
)


def exporta_datos(datos):
    output = BytesIO()
    datos.write_excel(workbook=output, autofilter=False)
    return output.getvalue()


resumen_sin_registro = html.Div(
        dbc.Col([
            dbc.Alert('No hay registros que visualizar', color='primary', style={'textAlign': 'center', 'marginTop': '25px'}),
        ],
        width={"size": 8, "offset": 2},
        ),
    )


def resumen(datos, modo):
    if datos == []:
        return resumen_sin_registro
    else:
        return (
            html.Div([
                vista_resumen(),
                modo_resumen(datos, modo),
                boton_descarga,
            ])
        )


#### Ingreso de infromación
# función que crea botones de radio

def crea_radio(ops, id_radio, ini=None, horizontal=False):
    return dcc.RadioItems(
        options=opciones(ops),
        value=ini if ini else None,
        id=id_radio,
        inline=horizontal,
        className='radio-gral',
        labelClassName='radio-label',
        inputClassName='radio-input',
        style = {'width': 'auto', 'display': 'block'},
    )


# funcion que crea fecha
def crea_fecha(ini, fin, id_dia, id_mes, id_ano, fecha=None, nt=None):
    if fecha is not None:
        val_dia, val_mes, val_ano = (eval(f'{nt}.{i}') for i in [f'dia_{fecha}', f'mes_{fecha}', f'ano_{fecha}'])
        hid = False
    else:
        val_dia, val_mes, val_ano = (None, None, None)
        hid = 'hidden'

    return html.Div([
        html.Div([
            dcc.Dropdown(dias, id=id_dia, value=val_dia, placeholder='Día', optionHeight=28),
            html.P(f'({str(val_dia or "")})', hidden=hid, style={'color': 'red'}),
            ], style={'width': '26%'},
        ),
        html.Div([
            dcc.Dropdown(opciones(meses), id=id_mes, value=val_mes, placeholder='Mes', optionHeight=28),
            html.P(f'({meses.get(val_mes, "")})', hidden=hid, style={'color': 'red'}),
            ], style={'width': '44%'},
        ),
        html.Div([
            dcc.Dropdown([x for x in range(ini, fin+1)], id=id_ano, value=val_ano, placeholder='Año', optionHeight=28),  
            html.P(f'({str(val_ano or "")})', hidden=hid, style={'color': 'red'}),
            ], style={'width': '30%'},
        ),
    ], style={'display': 'flex'})


ing_documento = dbc.Row([
    dbc.Col([
        dbc.Label('Tipo de documento:', className='label-ant'),
        crea_radio(documento, 'sel-documento', horizontal=True) # no acepta selección inicial
    ],
    width={'offset': 1},
    style={'width': '16%'},
    ),
    dbc.Col([
        dbc.Label('Número del documento:', className='label-ant'),
        dbc.Input(id='inp-numero', placeholder='Ingrese el número aquí', type='number'),
    ],
    style={'width': '19%'},
    ),
    dbc.Col([
        dbc.Label('Dígito verificador:', className='label-ant'),
        html.Div(
            dcc.Dropdown(digito, id='sel-dv', placeholder='DV', optionHeight=28),
            style={'width': '50%'},
        ),        
    ],
    style={'width': '14%'},
    ),
],
className='row-ajuste',
)

ing_postulante = dbc.Row([
    dbc.Col([
        html.Label('Apellido 1: ', className='label-ant'),
        dbc.Input(id='inp-apellido1', className='input-gral', placeholder='Apellido', type='text'),
    ],
    width={'offset': 1},
    style={'width': '17%'},
    ),
    dbc.Col([
        html.Label('Apellido 2: ', className='label-ant'),
        dbc.Input(id='inp-apellido2', className='input-gral', placeholder='Apellido', type='text'),
    ],
    style={'width': '17%'},
    ),
    dbc.Col([
        html.Label('Nombres: ', className='label-ant'),
        dbc.Input(id='inp-nombres', className='input-gral', placeholder='Nombres', type='text'),
    ], 
    style={'width': '34%'},
    ),
    dbc.Col([
        html.Label('Sexo: ', className='label-ant'),
        crea_radio(sexo, 'sel-sexo', horizontal=True),
    ],
    style={'width': '23%'},
    ),
],
className='row-ajuste',
)

ing_postulante2 = dbc.Row([
    dbc.Col([
        html.Label('Fecha de nacimiento:', className='label-ant'),
        crea_fecha(1950, 2010, 'dia-nac', 'mes-nac', 'ano-nac'),
    ],
    width={'offset': 1},
    style={'width': '27%'},
    ),
    dbc.Col([
        html.Label('Nacionalidad:', className='label-ant'),
        html.Div(
            dcc.Dropdown(opciones(pais), id='nac-postl', placeholder='Nacionalidad', optionHeight=28, style={'marginRight': '8px'}),
        ),
    ],
    style={'width': '17%'},
    ),
    dbc.Col([
        html.Label('País de origen:', className='label-ant'),
        html.Div(
            dcc.Dropdown(opciones(pais), id='pais-postl', placeholder='País', optionHeight=28, style={'marginRight': '8px'}),
        ),
    ],
    style={'width': '17%'},
    ),
    dbc.Col([
        html.Label('Tipo de residencia:', className='label-ant'),
        html.Div(
            dcc.Dropdown(opciones(tipo_residencia), id='resid-postl', placeholder='Tipo residencia', optionHeight=28, style={'marginRight': '8px'}),
        ),
    ],
    style={'width': '30%'},
    ),
],
className='row-ajuste',
)

ing_programa = dbc.Row([
    dbc.Col([
        dbc.Label('Nombre programa:', className='label-ant'),
        dbc.Textarea(id='nom-prog', placeholder="Nombre del programa", rows=1),
    ], width={'offset': 1, 'size': 8}
    ),
    dbc.Col([
        dbc.Label('Código programa:', className='label-ant'),
        dbc.Input(id='cod-prog', placeholder='Código del programa', type='text'),
    ],
    style={'width': '20%'},
    ),
],
align='center',
className='row-ajuste',
)

ing_programa2 = dbc.Row([
    dbc.Col([
        dbc.Label('Tipo programa:', className='label-ant'),
        crea_radio(tipo_programa, 'sel-tipo-prog'),
    ],
    width={'offset': 1},
    style={'width': '20%'},
    ),
    dbc.Col([
        dbc.Label('Especificación programa:', className='label-ant'),
        dbc.Col(
            dbc.Textarea(id='esp-prog', placeholder="Especificación del programa", rows=3),
        ),
    ],
    style={'width': '55%'},
    ),
    dbc.Col([
        dbc.Label('Duración programa:', className='label-ant'),
        html.Div([
            dbc.Input(id='ing-durac-prog', className='input-gral input-durac', placeholder='Meses', type='number'),
            dbc.Label('meses'),
        ]),
    ],
    style={'width': '16%'}
    ),
],
className='row-ajuste',
)

ing_programa3 = dbc.Row([
    dbc.Col([
        dbc.Label('Fecha de inicio:', className='label-ant'),
        crea_fecha(2024, 2028, 'dia-ini', 'mes-ini', 'ano-ini'),
    ],
    width={'offset': 1},
    style={'width': '26%'}
    ),
    dbc.Col([
        dbc.Label('Fecha de término:', className='label-ant'),
        crea_fecha(2024, 2028, 'dia-ter', 'mes-ter', 'ano-ter'),
    ],
    style={'width': '26%'}
    ),
    dbc.Col([
        dbc.Label('Jornada:', className='label-ant'),
        crea_radio(jornada, 'sel-jornada', horizontal=True)
    ],
    style={'width': '23%'}
    ),
    dbc.Col([
        dbc.Label('Comuna programa:', className='label-ant'),
        html.Div(
            dcc.Dropdown(comunas, id='sel-comuna', placeholder='Comuna', optionHeight=28, style={'marginRight': '8px'}),
        ),
    ],
    style={'width': '15%'}
    ), 
],
className='row-ajuste',
)

ing_universidad = dbc.Row([
    dbc.Col([
        dbc.Label('Universidad de origen:', className='label-ant'),
        dbc.Input(id='ing-univ-ori', className='input-gral', placeholder='Universidad de origen', type='text'),
    ],
    width={'offset': 1},
    style={'width': '50%'}
    ), 
    dbc.Col([
        dbc.Label('País:', className='label-ant'),
        html.Div(
            dcc.Dropdown(opciones(pais), id='pais-univ', placeholder='País', optionHeight=28, style={'marginRight': '8px'}),
        ),        
    ],
    style={'width': '17%'}
    ), 
    dbc.Col([
        dbc.Label('Existe convenio:', className='label-ant'),
        crea_radio(convenio, 'sel-convenio', horizontal=True)
    ],
    style={'width': '15%'}
    ), 
],
className='row-ajuste',
)

ing_seguimiento = dbc.Row([
    dbc.Col([
        dbc.Label('Fecha de postulación:', className='label-ant'),
        crea_fecha(2024, 2025, 'dia-pos', 'mes-pos', 'ano-pos'),
    ],
    width={'offset': 1},
    style={'width': '26%'}
    ),
    dbc.Col([
        dbc.Label('Condición de la información:', className='label-ant'),
        crea_radio(informacion, 'sel-infor', horizontal=True)
    ],
    style={'width': '20%'}
    ), 
    dbc.Col([
        dbc.Label('Estatus postulación:', className='label-ant'),
        crea_radio(estatus, 'sel-status', horizontal=True)
    ],
    style={'width': '25%'}
    ), 
],
className='row-ajuste',
)


boton_ingresa = dbc.Row([
    html.Hr(style={'marginTop': 5, 'marginBottom': 5, 'border': None, 'borderWidth': '1px', 'width': '100%', 'color': azul, 'opacity': 1}),
    html.Button('Limpiar campos', id='btn-limp-inf', n_clicks=0, className='btn btn-outline-primary', style={'width': '17%'}),
    html.Button('Ingreso de información', id='btn-ing-inf', n_clicks=0, disabled=True, className='btn btn-outline-primary', style={'width': '17%', 'marginLeft': '10px'}),
],
    justify='end',
    style={'margin': '0px 0px 35px'}
)


# modales

ingreso_correcto = html.Div(
    dbc.Modal(
        [
            dbc.ModalHeader(html.H4('Ingreso de información')),
            dbc.ModalBody(html.P('Se ha ingresado satisfactoriamente un registro a la base de datos.')),
            dbc.ModalFooter(dbc.Button('Cerrar', id='btn-mod-ingreso', n_clicks=0, outline=True, color='primary', style={'width': '30%'})),
        ],
        id='modal-ing-correcto',
    ),
)

registro_duplicado = html.Div(
    dbc.Modal(
        [
            dbc.ModalHeader(html.H4('Registro duplicado')),
            dbc.ModalBody(html.P('Existe un registro idéntico en la base de datos, por lo que la información no será ingresada.')),
            dbc.ModalFooter(dbc.Button('Cerrar', id='btn-mod-reg-dup', n_clicks=0, outline=True, color='primary', style={'width': '30%'})),
        ],
        id='modal-reg-duplicado',
    ),
)

actualiza_registro = html.Div(
    dbc.Modal(
        [
            dbc.ModalHeader(html.H4('Actualización de registro')),
            dbc.ModalBody(html.P('La información ingresada reemplazará a información existente en la base de datos.')),
            dbc.ModalFooter(
                dbc.Row([
                    html.Button('Cancelar', id='btn-no-actualiza', n_clicks=0, className='btn btn-outline-primary', style={'width': '80%'}),
                    html.Button('Aceptar', id='btn-si-actualiza', n_clicks=0, className='btn btn-outline-primary', style={'width': '80%', 'marginTop': '5px'}),
                ], justify='end'),
            ),
        ],
        id='modal-actualiza-reg',
    ),
)

lista_id_ing = ['sel-documento', 'inp-numero', 'sel-dv', 'inp-apellido1', 'inp-apellido2', 'inp-nombres', 'sel-sexo', 'dia-nac', 'mes-nac', 'ano-nac', 'nac-postl',
                'pais-postl', 'resid-postl', 'nom-prog', 'cod-prog', 'sel-tipo-prog', 'esp-prog', 'ing-durac-prog', 'dia-ini', 'mes-ini', 'ano-ini', 'dia-ter', 'mes-ter',
                'ano-ter', 'sel-jornada', 'sel-comuna', 'ing-univ-ori', 'pais-univ', 'sel-convenio', 'dia-pos', 'mes-pos', 'ano-pos', 'sel-infor', 'sel-status']


def ingresa():
    return (
        dbc.Form(
            dbc.Row([
                html.H5('Documento'),
                ing_documento,
                html.H5('Postulante'),
                ing_postulante,
                ing_postulante2,
                html.H5('Programa'),
                ing_programa,
                ing_programa2,
                ing_programa3,
                html.H5('Universidad Origen'),
                ing_universidad,
                html.H5('Seguimiento'),
                ing_seguimiento,
                boton_ingresa,
                ingreso_correcto,
                registro_duplicado,
                actualiza_registro,
            ])
        )
    )


#### Ingresa Excel
subir_excel = html.Div([
    dbc.Row(
        dbc.Col(
            html.Div(
                dcc.Upload(
                    id='upload-excel',
                    children=html.Div([
                        'Mueva el archivo hasta este recuadro o ',
                        html.A('Selecciónelo con el navegador', style={'color': azul})
                    ],
                    className='up-excel',
                    ),
                    multiple=False
                ),
                id='wrap-upload',
            ), width={'size': 8, 'offset': 2}
        ),
    ),
    dbc.Row(
        html.Div(id='output-data-upload'),
    )
])


def ingresa_excel():
    return subir_excel

#### Modificación/Eliminación de información
columns_def_modifica = [
    {'field': 'id', 'hide': True},
    {'field': 'documento', 'headerName': 'Documento', 'width': 120},
    {'field': 'numero', 'headerName': 'Número', 'width': 150},
    {'field': 'dv', 'headerName': 'DV', 'width': 60},
    {'field': 'apellido_1', 'headerName': 'Apellido 1'},
    {'field': 'apellido_2', 'headerName': 'Apellido 2'},
    {'field': 'nombres', 'headerName': 'Nombres', 'width': 150},
    {'field': 'sexo', 'headerName': 'Sexo'},
    {'field': 'fecha_nac', 'headerName': 'Fecha de nacimiento'},
    {'field': 'pais', 'headerName': 'Nacionalidad'},
    {'field': 'nacional', 'headerName': 'País'},
    {'field': 'tipo_residencia', 'headerName': 'Tipo de residencia'},
    {'field': 'programa', 'headerName': 'Nombre programa'},
    {'field': 'cod_programa', 'headerName': 'Código del programa'},
    {'field': 'tipo_programa', 'headerName': 'Tipo programa'},
    {'field': 'esp_programa', 'headerName': 'Especificación programa'},
    {'field': 'duracion', 'headerName': 'Duración'},
    {'field': 'inicio_programa', 'headerName': 'Fecha de inicio'},
    {'field': 'termino_programa', 'headerName': 'Fecha de término'},
    {'field': 'jornada', 'headerName': 'Jornada'},
    {'field': 'comuna', 'headerName': 'Comuna'},
    {'field': 'universidad_origen', 'headerName': 'Universidad de origen'},
    {'field': 'pais_universidad', 'headerName': 'País de la universidad'},
    {'field': 'convenio', 'headerName': 'Existe convenio'},
    {'field': 'fecha_postulacion', 'headerName': 'Fecha postulación'},
    {'field': 'condicion_inf', 'headerName': 'Condición información'},
    {'field': 'estatus', 'headerName': 'Estatus postulación'},
    {'field': 'universidad', 'hide': True},
]


def op_modificar(datos):
    if datos == []:
        return modifica_sin_registro
    else:
        df = back_excel(
            pl.DataFrame(datos)
            .cast(sch_base_loc)
            .select(['id'] + lista_excel + ['universidad'])
        )
        return html.Div(
            dbc.Row([
                dbc.Col(
                    dag.AgGrid(
                        id='selector-mod-eli',
                        rowData=df.to_dicts(),
                        columnDefs=columns_def_modifica,
                        dashGridOptions = {
                            'rowSelection': 'single',
                            'domLayout': 'autoHeight',
                        },
                        defaultColDef = {'editable': False}
                    ),
                    style={'width': '85%'}
                ),
                dbc.Col([
                    html.Button('Modificar', id='btn-modifica', n_clicks=0, className='btn btn-outline-primary', disabled=True, style={'width': '100%'}),
                    html.Button('Eliminar', id='btn-elimina', n_clicks=0, className='btn btn-outline-primary', disabled=True, style={'width': '100%', 'marginTop': '5px'}),
                ],
                style={'width': '15%'}
                ),
            ])
        )


modifica_sin_registro = html.Div(
        dbc.Col([
            dbc.Alert('No hay registros para modificar/eliminar', color='primary', style={'textAlign': 'center', 'marginTop': '25px'}),
        ],
        width={"size": 8, "offset": 2},
        ),
    )


def modifica(datos):
    if datos == []:
        return modifica_sin_registro
    else:
        return (
            html.Div([
                html.Div([
                    html.H4('Registros susceptibles de ser modificados/eliminados', style={'marginBottom': '0.8rem'}),
                    html.Div(id='contenido-mod'), # contenedor 1
                ], id='cont-mod1'),
                html.Div(id='pantalla-mod'), # contenedor 2
            ])
        )


def fn_mod_documento(nt):
    return dbc.Row([
        dbc.Col([
            dbc.Label('Tipo de documento:', className='label-ant2'),
            crea_radio(documento, 'mod-documento', ini=nt.documento, horizontal=True),
            html.P(f'({documento.get(nt.documento, "")})', style={'color': 'red'}),
        ],
        width={'offset': 1},
        style={'width': '16%'},
        ),
        dbc.Col([
            dbc.Label('Dígito verificador:', className='label-ant2'),
            html.Div(
                dcc.Dropdown(digito, id='mod-dv', value=nt.dv, placeholder='DV', optionHeight=28),
                style={'width': '50%'},
            ),        
            html.P(f'({str(nt.dv or "")})', style={'color': 'red'}),
        ],
        style={'width': '14%'},
        ),
    ],
    className='row-ajuste',
    )


def fn_mod_postulante(nt):
    return dbc.Row([
        dbc.Col([
            html.Label('Apellido 1: ', className='label-ant2'),
            dbc.Input(id='mod-apellido1', className='input-gral', value=nt.apellido_1, placeholder='Apellido', type='text'),
            html.P(f'({str(nt.apellido_1 or "")})', style={'color': 'red'}),
        ],
        width={'offset': 1},
        style={'width': '17%'},
        ),
        dbc.Col([
            html.Label('Apellido 2: ', className='label-ant2'),
            dbc.Input(id='mod-apellido2', className='input-gral', value=nt.apellido_2, placeholder='Apellido', type='text'),
            html.P(f'({str(nt.apellido_2 or "")})', style={'color': 'red'}),
        ],
        style={'width': '17%'},
        ),
        dbc.Col([
            html.Label('Nombres: ', className='label-ant2'),
            dbc.Input(id='mod-nombres', className='input-gral', value=nt.nombres, placeholder='Nombres', type='text'),
            html.P(f'({str(nt.nombres or "")})', style={'color': 'red'}),
        ],
        style={'width': '34%'},
        ),
        dbc.Col([
            html.Label('Sexo: ', className='label-ant2'),
            crea_radio(sexo, 'mod-sexo', ini=nt.sexo, horizontal=True),
            html.P(f'({sexo.get(nt.sexo, "")})', style={'color': 'red'}),
        ],
        style={'width': '23%'},
        ),
    ],
    className='row-ajuste',
    )


def fn_mod_postulante2(nt):
    return dbc.Row([
        dbc.Col([
            html.Label('Fecha de nacimiento:', className='label-ant2'),
            crea_fecha(1950, 2010, 'mod-dia-nac', 'mod-mes-nac', 'mod-ano-nac', fecha='nac', nt=nt),
        ],
        width={'offset': 1},
        style={'width': '27%'},
        ),
        dbc.Col([
            html.Label('Nacionalidad:', className='label-ant2'),
            html.Div(
                dcc.Dropdown(opciones(pais), id='mod-nac-postl', value=nt.nacional, placeholder='Nacionalidad', optionHeight=28, style={'marginRight': '8px'}),
            ),
            html.P(f'({pais.get(nt.nacional, "")})', style={'color': 'red'}),
        ],
        style={'width': '17%'},
        ),
        dbc.Col([
            html.Label('País de origen:', className='label-ant2'),
            html.Div(
                dcc.Dropdown(opciones(pais), id='mod-pais-postl', value=nt.pais, placeholder='País', optionHeight=28, style={'marginRight': '8px'}),
            ),
            html.P(f'({pais.get(nt.pais, "")})', style={'color': 'red'}),
        ],
        style={'width': '17%'},
        ),
        dbc.Col([
            html.Label('Tipo de residencia:', className='label-ant2'),
            html.Div(
                dcc.Dropdown(opciones(tipo_residencia), id='mod-resid-postl', value=nt.tipo_residencia, placeholder='Tipo residencia', optionHeight=28, style={'marginRight': '8px'}),
            ),
            html.P(f'({tipo_residencia.get(nt.tipo_residencia, "")})', style={'color': 'red'}),
        ],
        style={'width': '30%'},
        ),
    ],
    className='row-ajuste',
    )


def fn_mod_programa(nt):
    return dbc.Row([
        dbc.Col([
            dbc.Label('Código programa:', className='label-ant2'),
            dbc.Input(id='mod-cod-prog', value=nt.cod_programa, placeholder='Código del programa', type='text'),
            html.P(f'({str(nt.cod_programa or "")})', style={'color': 'red'}),
        ],
        width={'offset': 1},
        style={'width': '20%'},
        ),
    ],
    align='center',
    className='row-ajuste',
    )



def fn_mod_programa2(nt):
    return dbc.Row([
        dbc.Col([
            dbc.Label('Tipo programa:', className='label-ant2'),
            crea_radio(tipo_programa, 'mod-tipo-prog', ini=nt.tipo_programa),
            html.P(f'({tipo_programa.get(nt.tipo_programa, "")})', style={'color': 'red'}),
        ],
        width={'offset': 1},
        style={'width': '20%'},
        ),
        dbc.Col([
            dbc.Label('Especificación programa:', className='label-ant2'),
            dbc.Col(
                dbc.Textarea(id='mod-esp-prog', value=nt.esp_programa, placeholder="Especificación del programa", rows=3),
            ),
            html.P(f'({str(nt.esp_programa or "")})', style={'color': 'red'}),
        ],
        style={'width': '55%'},
        ),
        dbc.Col([
            dbc.Label('Duración programa:', className='label-ant2'),
            html.Div([
                dbc.Input(id='mod-durac-prog', className='input-gral input-durac', value=nt.duracion, placeholder='Meses', type='number'),
                dbc.Label('meses'),
            ]),
            html.P(f'({str(nt.duracion or "")})', style={'color': 'red'}),
        ],
        style={'width': '16%'}
        ),
    ],
    className='row-ajuste',
    )


def fn_mod_programa3(nt):
    return dbc.Row([
        dbc.Col([
            dbc.Label('Fecha de inicio:', className='label-ant2'),
            crea_fecha(2024, 2028, 'mod-dia-ini', 'mod-mes-ini', 'mod-ano-ini', fecha='ini', nt=nt),
        ],
        width={'offset': 1},
        style={'width': '26%'}
        ),
        dbc.Col([
            dbc.Label('Fecha de término:', className='label-ant2'),
            crea_fecha(2024, 2028, 'mod-dia-ter', 'mod-mes-ter', 'mod-ano-ter', fecha='ter', nt=nt),
        ],
        style={'width': '26%'}
        ),
        dbc.Col([
            dbc.Label('Jornada:', className='label-ant2'),
            crea_radio(jornada, 'mod-jornada', ini=nt.jornada, horizontal=True),
            html.P(f'({jornada.get(nt.jornada, "")})', style={'color': 'red'}),
        ],
        style={'width': '23%'}
        ),
        dbc.Col([
            dbc.Label('Comuna programa:', className='label-ant2'),
            html.Div(
                dcc.Dropdown(comunas, id='mod-comuna', value=nt.comuna, placeholder='Comuna', optionHeight=28, style={'marginRight': '8px'}),
            ),
            html.P(f'({str(nt.comuna or "")})', style={'color': 'red'}),
        ],
        style={'width': '15%'}
        ), 
    ],
    className='row-ajuste',
    )


def fn_mod_universidad(nt):
    return dbc.Row([
        dbc.Col([
            dbc.Label('Universidad de origen:', className='label-ant2'),
            dbc.Input(id='mod-univ-ori', className='input-gral', value=nt.universidad_origen, placeholder='Universidad de origen', type='text'),
            html.P(f'({str(nt.universidad_origen or "")})', style={'color': 'red'}),
        ],
        width={'offset': 1},
        style={'width': '50%'}
        ), 
        dbc.Col([
            dbc.Label('País:', className='label-ant2'),
            html.Div(
                dcc.Dropdown(opciones(pais), id='mod-pais-univ', value=nt.pais_universidad, placeholder='País', optionHeight=28, style={'marginRight': '8px'}),
            ),        
            html.P(f'({pais.get(nt.pais_universidad, "")})', style={'color': 'red'}),
        ],
        style={'width': '17%'}
        ), 
        dbc.Col([
            dbc.Label('Existe convenio:', className='label-ant2'),
            crea_radio(convenio, 'mod-convenio', ini=nt.convenio, horizontal=True),
            html.P(f'({convenio.get(nt.convenio, "")})', style={'color': 'red'}),
        ],
        style={'width': '15%'}
        ), 
    ],
    className='row-ajuste',
    )


def fn_mod_seguimiento(nt):
    return dbc.Row([
    dbc.Col([
            dbc.Label('Fecha de postulación:', className='label-ant2'),
            crea_fecha(2024, 2025, 'mod-dia-pos', 'mod-mes-pos', 'mod-ano-pos', fecha='pos', nt=nt),
        ],
        width={'offset': 1},
        style={'width': '26%'}
        ),
        dbc.Col([
            dbc.Label('Condición de la información:', className='label-ant2'),
            crea_radio(informacion, 'mod-infor', ini=nt.condicion_inf, horizontal=True),
            html.P(f'({informacion.get(nt.condicion_inf, "")})', style={'color': 'red'}),
        ],
        style={'width': '20%'}
        ), 
        dbc.Col([
            dbc.Label('Estatus postulación:', className='label-ant2'),
            crea_radio(estatus, 'mod-status', ini=nt.estatus, horizontal=True),
            html.P(f'({estatus.get(nt.estatus, "")})', style={'color': 'red'}),
        ],
        style={'width': '25%'}
        ), 
    ],
    className='row-ajuste',
    )


boton_modifica = dbc.Row([
    html.Hr(style={'marginTop': 5, 'marginBottom': 5, 'border': None, 'borderWidth': '1px', 'width': '100%', 'color': azul, 'opacity': 1}),
    html.Button('Salir sin guardar', id='btn-canc-mod', n_clicks=0, className='btn btn-outline-primary', style={'width': '17%'}),
    html.Button('Aplicar cambios', id='btn-apli-mod', n_clicks=0, disabled=True, className='btn btn-outline-primary', style={'width': '17%', 'marginLeft': '10px'}),
],
    justify='end',
    style={'margin': '0px 0px 35px'}
)

lista_id_mod = ['mod-documento', 'mod-dv', 'mod-apellido1', 'mod-apellido2', 'mod-nombres', 'mod-sexo', 'mod-dia-nac', 'mod-mes-nac', 'mod-ano-nac', 'mod-nac-postl',
               'mod-pais-postl', 'mod-resid-postl', 'mod-cod-prog', 'mod-tipo-prog', 'mod-esp-prog', 'mod-durac-prog', 'mod-dia-ini', 'mod-mes-ini', 'mod-ano-ini',
               'mod-dia-ter', 'mod-mes-ter', 'mod-ano-ter', 'mod-jornada', 'mod-comuna', 'mod-univ-ori', 'mod-pais-univ', 'mod-convenio', 'mod-dia-pos', 'mod-mes-pos',
               'mod-ano-pos', 'mod-infor', 'mod-status']


def fn_inalterables(nt):
    return html.Div([
        dbc.Col([
            dbc.Row([
                dbc.Label('Número de documento:', className='label-otro', style={'background': '#FCECFC'}),
                html.P(f'{nt.numero}', style={'width': '83%', 'margin': '8px 0px'}),
            ], align='center'),
            dbc.Row([
                dbc.Label('Nombre del programa:', className='label-otro', style={'background': '#FCECFC'}),
                html.P(f'{nt.programa}', style={'width': '83%', 'margin': '8px 0px'}),
            ], align='center'),
        ],
        width={'offset': 1},
        )
    ], style={'marginBottom': '10px'})


color_h = '#A615A6'

def fn_modifica(nt):
    return html.Div([
        html.H5('Campos inalterables', style={'color': color_h}),
        fn_inalterables(nt),
        html.H5('Documento', style={'color': color_h}),
        fn_mod_documento(nt),
        html.H5('Postulante', style={'color': color_h}),
        fn_mod_postulante(nt),
        fn_mod_postulante2(nt),
        html.H5('Programa', style={'color': color_h}),
        fn_mod_programa(nt),
        fn_mod_programa2(nt),
        fn_mod_programa3(nt),
        html.H5('Universidad Origen', style={'color': color_h}),
        fn_mod_universidad(nt),
        html.H5('Seguimiento', style={'color': color_h}),
        fn_mod_seguimiento(nt),
        boton_modifica,
    ])


#### Eliminación de información
def elimina():
    return (
        html.Div(
            html.H3('Eliminación de registro')
        )
    )


#### Pie de página
def form_footer():
    return html.Div(
        dbc.Row(
            dbc.Col(
                html.Footer(
                    ['2025:  Corporación de Universidades Privadas'], className='ob-footer',
                )
            )
        ),
        className='footer-strip'
    )


#### Contenido
lista_disparador = ['tab-resumen', 'tab-ingresa_1', 'tab-ingresa_2', 'tab-modifica']
lista_paginas = ['Resumen', 'Ingreso de información', 'Información desde archivo Excel', 'Modificación/eliminación de información']

dic_disparador = dict(zip(lista_disparador, [resumen, ingresa, ingresa_excel, modifica]))
dic_ubicacion = dict(zip(lista_disparador, [f': {x}' for x in lista_paginas]))

contenido = html.Div(id='contenido')

parametros_iniciales = {
    'usuario': 0,
}


#### Layout
# layout de la aplicación

def serve_layout():
    return dbc.Container([
        html.Div([
            acceso,
            encabezado,
            muestra_usuario(),
            dd_menu,
            linea,
            contenido,
        ]),
        form_footer(),

        dcc.Store(id='datos', data=[]),
        dcc.Store(id='provisional', data=[]),
        dcc.Store(id='parametros', data=parametros_iniciales),

    ], className='layout-total')


### Aplicación

app = DashProxy(__name__, transforms=[MultiplexerTransform()], external_stylesheets=[dbc.themes.CERULEAN])

app.config.suppress_callback_exceptions = True

app.layout = serve_layout

#
# Callbacks
#

# Acceso
@app.callback(
    Output('acceso', 'hidden'),
    Output('sp-usuario', 'children'),
    Output('encabezado', 'hidden'),
    Output('muestra-usuario', 'hidden'),
    Output('menu', 'hidden'),
    Output('linea', 'hidden'),
    Output('contenido', 'children'),
    Output('inp-pw', 'value'),
    Output('datos', 'data'),
    Output('parametros', 'data'),
    Input('btn-ingresar', 'n_clicks'),
    State('sel-u', 'value'),
    State('inp-pw', 'value'),
    prevent_initial_call=True,
)
def autenticacion(click, usuario, pw):
    if click == 0:
        raise PreventUpdate
    else:
        if pw == os.environ.get('U'+str(usuario)):
            datos = lectura_conv(usuario)[1]
            return 'hidden', universidades[usuario], False, False, False, False, resumen(datos, 1), [], datos, {'usuario': usuario}
        else:
            return dash.no_update, dash.no_update, dash.no_update, dash.no_update, dash.no_update, dash.no_update, dash.no_update, [], dash.no_update, dash.no_update


# Navegación
@app.callback(
    Output('contenido', 'children'),
    Output('tab-resumen', 'active'),
    Output('tab-ingresa_1', 'active'),
    Output('tab-ingresa_2', 'active'),
    Output('tab-modifica', 'active'),
    Output('ubica', 'children'),
    Input('tab-resumen', 'n_clicks'),
    Input('tab-ingresa_1', 'n_clicks'),
    Input('tab-ingresa_2', 'n_clicks'),
    Input('tab-modifica', 'n_clicks'),
    State('datos', 'data'),
    prevent_initial_call=True,
)
def selecciona_pagina(p1, p2, p3, p4, datos):
    disparador = dash.ctx.triggered_id
    ret = [x == disparador for x in lista_disparador]

    if disparador == 'tab-resumen':
        return resumen(datos, 1), *ret, dic_ubicacion[disparador]
    elif disparador == 'tab-ingresa_1':
        return ingresa(), *ret, dic_ubicacion[disparador]
    elif disparador == 'tab-ingresa_2':
        return ingresa_excel(), *ret, dic_ubicacion[disparador]
    elif disparador == 'tab-modifica':
        return modifica(datos), *ret, dic_ubicacion[disparador]


# Ingreso desde Archivo Excel
# agregar datos a agregar a la base
@app.callback(
    Output('output-data-upload', 'children'),
    Output('provisional', 'data'),
    Output('wrap-upload', 'hidden'),
    Input('upload-excel', 'contents'),
    State('upload-excel', 'filename'),
    State('parametros', 'data'),
    prevent_initial_call=True,
)
def ingreso_desde_excel(contenido, nombre, param):  # retornar el contenido del aggrid y la data a ser ingresada
    if contenido is not None:
        grid, dat_algo = parse_excel(contenido, nombre, param['usuario'])
        return grid, dat_algo, 'hidden'
    else:
        raise PreventUpdate


# Botón que restituye página de ingreso Excel
@app.callback(
    Output('contenido', 'children'),
    Input('btn-restituye', 'n_clicks'),
    prevent_initial_call=True,
)
def restituye_upload_excel(click):
    if click == 0:
        raise PreventUpdate
    else:
        return ingresa_excel()


# Segundo botón
@app.callback(
    Output('contenido', 'children'),
    Input('btn-canc-inf-ex', 'n_clicks'),
    prevent_initial_call=True,
)
def restituye_upload_excel2(click):
    if click == 0:
        raise PreventUpdate
    else:
        return ingresa_excel()


# NN
@app.callback(
    Output('contenido', 'children'),
    Output('datos', 'data'),
    Input('btn-ing-inf-ex', 'n_clicks'),
    State('provisional', 'data'),
    State('parametros', 'data'),
    State('sel-registros', 'value'),
    prevent_initial_call=True,
)
def ingreso_desde_excel(click, datos, param, todos):
    if click == 0:
        raise PreventUpdate
    else:
        agrega_inf_excel(datos, todos)
        return ingresa_excel(), lectura_conv(param['usuario'])[1]


# Selección del modo de visualización del resumen
@app.callback(
    Output('modo-seleccionado', 'children'),
    Input('selector-modo', 'value'),
    State('datos', 'data'),
    prevent_initial_call=True,
)
def cambia_modo_visualizacion(modo, data):
    return modo_resumen(data, modo)


# Ingreso de la información desde la app: general
@app.callback(
    Output('datos', 'data'),
    Output('modal-ing-correcto', 'is_open'),
    Output('modal-reg-duplicado', 'is_open'),
    Output('modal-actualiza-reg', 'is_open'),
    Output('provisional', 'data'),
    Input('btn-ing-inf', 'n_clicks'),
    State('parametros', 'data'),
    [State(i, 'value') for i in lista_id_ing],
    prevent_initial_call=True,
)
def ingreso_desde_aplicacion(click, param, *lista_app):
    if click == 0:
        raise PreventUpdate
    else:
        tupla = Registro_app(*lista_app)
        datos, modal1, modal2, modal3, provision = parse_app(tupla, param['usuario'])
        return datos, modal1, modal2, modal3, provision


# Activa posibilidad de guardar registro
@app.callback(
    Output('btn-ing-inf', 'disabled'),
    Input('inp-numero', 'value'),
    Input('nom-prog', 'value'),    
)
def activa_ingreso_datos(numero, prog):
    if (numero != None) & (prog != None) & (prog != ''):
        return False
    else:
        return True


# Cierra modal cuando es duplicado
@app.callback(
    Output('modal-reg-duplicado', 'is_open'),
    Input('btn-mod-reg-dup', 'n_clicks'),
    prevent_initial_call=True,
)
def cierra_modal_duplicado(click):
    if click == 0:
        raise PreventUpdate
    else:
        return False


# Cierra modal cuando ingresa nuevo
@app.callback(
    Output('modal-ing-correcto', 'is_open'),
    Input('btn-mod-ingreso', 'n_clicks'),
    prevent_initial_call=True,
)
def cierra_modal_ingreso(click):
    if click == 0:
        raise PreventUpdate
    else:
        return False


# Cierra modal cuando actualiza registro
@app.callback(
    Output('modal-actualiza-reg', 'is_open'),
    Output('datos', 'data'),
    Input('btn-no-actualiza', 'n_clicks'),
    Input('btn-si-actualiza', 'n_clicks'),
    State('provisional', 'data'),
    State('parametros', 'data'),
    prevent_initial_call=True,
)
def cierra_modal_actualiza(no, si, data, param):
    disparador = dash.ctx.triggered_id
    if disparador == 'btn-no-actualiza':
        return False, dash.no_update
    elif disparador == 'btn-si-actualiza':
        tup = Registro(**convierte_str_fecha(pl.DataFrame(data, schema_overrides=sch_base_loc)).rows(named=True)[0])
        modifica_visas(tup)
        return False, lectura_conv(param['usuario'])[1]


# Limpia campos de la ventana de ingreso
@app.callback(
    [Output(i, 'value') for i in lista_id_ing],
    Input('btn-limp-inf', 'n_clicks'),
    prevent_initial_call=True,
)
def limpia_campos_ingreso(click):
    if click == 0:
        raise PreventUpdate
    else:
        return None, *[None]*(len(lista_id_ing)-1)


# contenido de página de modificación
@app.callback(
    Output('contenido-mod', 'children'),
    Input('tab-modifica', 'n_clicks'),
    State('datos', 'data'),
)
def puebla_modifica(click, datos):
    return op_modificar(datos)


# botón elimina
@app.callback(
    Output('contenido-mod', 'children'),
    Output('datos', 'data'),
    Input('btn-elimina', 'n_clicks'),
    State('selector-mod-eli', 'selectedRows'),
    prevent_initial_call=True,
)
def elimina_seleccion(click, seleccion):
    if click == 0:
        raise PreventUpdate
    else:
        id = seleccion[0]['id']
        usuario = seleccion[0]['universidad']
        nuevos = elimina_visas(id, usuario)
        return op_modificar(nuevos), nuevos


# cambia condición de botones
@app.callback(
    Output('btn-modifica', 'disabled'),
    Output('btn-elimina', 'disabled'),
    Input('selector-mod-eli', 'selectedRows')
)
def activa_botones_mod_eli(seleccionado):
    if seleccionado:
        return False, False
    else:
        return True, True


# boton modifica
@app.callback(
    Output('pantalla-mod', 'children'),
    Output('cont-mod1', 'hidden'),
    Input('btn-modifica', 'n_clicks'),
    State('datos', 'data'),
    State('selector-mod-eli', 'selectedRows'),
    prevent_initial_call=True,
)
def modifica_seleccion(click, datos, seleccionado):
    if click == 0:
        raise PreventUpdate
    else:
        id_sel = seleccionado[0]['id']
        dic_sel = next(dic for dic in datos if dic['id'] == id_sel)
        nt = Registro(**dic_sel)
        return fn_modifica(nt), 'hidden'


# botón aplica modificaciones
@app.callback(
    Output('contenido', 'children'),
    Output('datos', 'data'),
    Output('selector-mod-eli', 'selectedRows'),
    Input('btn-apli-mod', 'n_clicks'),
    State('selector-mod-eli', 'selectedRows'),
    [State(i, 'value') for i in lista_id_mod],
    prevent_inital_call=True,
)
def aplica_modificaciones(click, seleccionado, *lista_mod):
    if click == 0:
        raise PreventUpdate
    else:
        tupla = Registro_mod(*lista_mod)
        dic_resto = {k: seleccionado[0][k] for k in ['id', 'numero', 'programa', 'universidad']}
        registro = convierte_mod(tupla, dic_resto)
        registro_sel = [{k: registro._asdict()[k] for k in lista_excel+['id', 'universidad']}]
        modifica_visas(registro)
        datos = lectura_conv(dic_resto['universidad'])[1]
        return modifica(datos), datos, registro_sel


# botón cancela modificaciones
@app.callback(
    Output('contenido', 'children'),
    Input('btn-canc-mod', 'n_clicks'),
    prevent_inital_call=True,
)
def cancela_modificaciones(click):
    if click == 0:
        raise PreventUpdate
    else:
        return modifica([1])


# activa botón que permite modificar registro
@app.callback(
    Output('btn-apli-mod', 'disabled'),
    State('selector-mod-eli', 'selectedRows'),
    State('datos', 'data'),
    [Input(i, 'value') for i in lista_id_mod],
)
def activa_boton_modifica(seleccionado, datos, *lista_mod):
    id_sel = seleccionado[0]['id']
    dic_sel = next(dic for dic in datos if dic['id'] == id_sel)
    dic = Registro_mod(*lista_mod)._asdict()
    dic_muestra = {k: v for k, v in dic_sel.items() if k in dic.keys()}
    return all([dic_muestra[k] == dic[k] for k in dic.keys()])


# descarga archivo de trabajo
@app.callback(
    Output('descarga-excel', 'data'),
    Input('btn-descarga-excel', 'n_clicks'),
    prevent_initial_call=True,
)
def func(n_clicks):
    return dcc.send_file('./data/VISA_ESTUDIANTES_INTERCAMBIO.xlsx')  # cambiar cuando corresponda


# exporta datos a excel
@app.callback(
    Output('exporta-excel', 'data'),
    Input('btn-exporta-excel', 'n_clicks'),
    State('datos', 'data'),
    State('selector-modo', 'value'),
    prevent_initial_call=True,
)
def exporta_datos_excel(_, datos, modo):
    df_loc = prepara_datos(datos, modo)
    df = exporta_datos(df_loc)
    return dcc.send_bytes(df, 'datos_visas.xlsx')


if __name__ == '__main__':
    app.run_server(debug=False)

